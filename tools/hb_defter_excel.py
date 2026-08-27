# -*- coding: utf-8 -*-
"""HB defter açıklamalarını, panelin toplu güncelleme Excel'i için hazırlar.

Neden Excel: HB'nin listing API'si açıklamayı hiç kabul etmiyor (stok/fiyat
dışındaki alanları sessizce yutuyor), `ticket-api/import` ucu da 200 dönüp
hiçbir şey yapmıyordu. Panelin **Toplu güncelleme → Ürün bilgisi güncelle**
yolu ise çalışıyor (27.08'de kargoya veriliş süresinde denendi ve tuttu).

Yöntem: panelden indirilen **mevcut** ürün bilgisi dışa aktarımı temel alınır;
yalnız "Ürün Açıklaması" hücresi değiştirilir. Böylece ürün adı ve görseller
olduğu gibi korunur — boş bırakılan hücrenin "değişmesin" mi "silinsin" mi
anlamına geldiğine güvenmek zorunda kalmayız.

    python tools/hb_defter_excel.py <disa_aktarim.xlsx> [<ikinci.xlsx> ...]
"""

from __future__ import annotations

import json
import pathlib
import sys

import openpyxl

KOK = pathlib.Path(__file__).resolve().parents[1]
sys.path.insert(0, str(KOK))

from tools.defter_aciklama import AILE, PAKET, metin, yaprak_ekle  # noqa: E402

# HB'deki SKU -> bizim stok kodumuz. Açıklama şablonu stok koduna göre
# hangi aile (iş dosyası / temrin) ve hangi paket olduğunu biliyor.
SKU_STOK = {
    "HBCV0000HPD033": "AESTJDFTR",   # tekli - iş dosyası (staj defteri)
    "HBCV0000HPD07X": "AETEMDEF",    # tekli - temrin
    "HBCV0000I2N6VA": "AEISDP10",
    "HBCV0000I2UXBM": "AEISDP20",
    "HBCV0000I1OFMT": "AEISD30",
    "HBCV0000I1OFS7": "AETMR10",
    "HBCV0000I2W9OC": "AETMRP20",
    "HBCV0000I2WEWJ": "AETMRP30",
}

KAYNAK = KOK / "tools" / "defter_tekli_kaynak.json"
CIKTI = KOK / "tools" / "hb-defter-aciklama.xlsx"


def tekli_metinler() -> dict:
    """Tekli açıklamaların ORİJİNAL hali.

    Canlı pazaryerinden okumak yanlış olur: TY zaten bu şablonla güncellendi,
    tekrar uygulamak paket bloklarını ikinci kez ekler.
    """
    return json.loads(KAYNAK.read_text(encoding="utf-8"))


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1

    kaynak = tekli_metinler()
    yazilan, atlanan = [], []
    hedef_wb = hedef_ws = None

    for yol in sys.argv[1:]:
        wb = openpyxl.load_workbook(yol)
        ws = wb[wb.sheetnames[0]]
        basliklar = [c.value for c in ws[1]]
        i_sku = basliklar.index("HB Ürün Kodu (SKU)")
        i_acik = basliklar.index("Ürün Açıklaması")

        if hedef_wb is None:
            hedef_wb, hedef_ws = wb, ws
            hedef_i_sku, hedef_i_acik = i_sku, i_acik
            satirlar = list(ws.iter_rows(min_row=2))
        else:
            # İkinci dosyadaki satırları ilkinin sonuna ekle.
            satirlar = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not any(r):
                    continue
                hedef_ws.append(list(r))
                satirlar.append(hedef_ws[hedef_ws.max_row])

        for satir in satirlar:
            sku = satir[hedef_i_sku].value
            if not sku:
                continue
            stok = SKU_STOK.get(str(sku).strip())
            if stok is None:
                atlanan.append(sku)
                continue
            aile = AILE[stok]
            if PAKET[stok] == "tekli":
                # HB'nin kendi tekli metni TY'dekinden zengin (2622 vs 2371).
                # Kullanicinin kurali: tekli aciklamalar iyi, KISALTMA. O yuzden
                # metni degistirmiyoruz, sadece yaprak/sayfa bilgisini ekliyoruz.
                yeni = yaprak_ekle(str(satir[hedef_i_acik].value or ""), aile)
            else:
                # Coklu paketlerde HB metni cok zayif (305-453 karakter). Burada
                # TY ile ayni sablon kullaniliyor - kullanici "TY'de neyse o"
                # dedigi icin iki pazaryeri birebir ayni metni gorsun.
                tekli = kaynak["AESTJDFTR"] if aile == "isdosyasi" else kaynak["AETEMDEF"]
                yeni = metin(stok, tekli)
            eski_uz = len(str(satir[hedef_i_acik].value or ""))
            satir[hedef_i_acik].value = yeni
            yazilan.append((stok, sku, eski_uz, len(yeni)))

    if atlanan:
        print("!! Eslesmeyen SKU (dokunulmadi):", ", ".join(map(str, atlanan)))

    print(f"{'stok kodu':12s} {'SKU':16s} {'eski':>6s} -> {'yeni':>6s}")
    for stok, sku, e, y in sorted(yazilan):
        print(f"{stok:12s} {sku:16s} {e:6d} -> {y:6d}")

    eksik = set(SKU_STOK.values()) - {s for s, *_ in yazilan}
    if eksik:
        print("\n!! Dosyalarda bulunamayan urunler:", ", ".join(sorted(eksik)))

    hedef_wb.save(CIKTI)
    print(f"\nyazildi: {CIKTI}  ({len(yazilan)} urun)")
    return 0


if __name__ == "__main__":
    sys.exit(main())

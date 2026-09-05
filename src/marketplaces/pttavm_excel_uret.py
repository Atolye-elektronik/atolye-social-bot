# -*- coding: utf-8 -*-
"""PttAVM 'Toplu Urun Yukleme' Excel'ini uretir (panel yolu; API self-entegrator bekliyor).

Fiyat: Trendyol ile esit net kar (pazarama ile ayni yontem).
  TY_net = P_ty*(1/1.2 - 0.13) - kargo_ty(P_ty) - 11
  P_ptt  = ceil( (TY_net + kargo_ptt(P_ptt, desi)) / (1/1.2 - komisyon) )
  kargo_ptt: 250 TL altinda alici oder (0); 250 ve ustunde desi tablosu (KDV dahil).
Komisyon oranlari PttAVM listesinde KDV DAHIL verilir.
"""
import json
import math
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from marketplaces.pazarama_urun_ekle import katalog, _duz_metin  # noqa: E402

KDV = 1.20
KOK = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# PttAVM yaprak kategori id + komisyon (KDV dahil, %)
KAT = {
    "bilim":      (1342, 0.15),   # Hobi & Oyuncak > Aktivite ve Egitici Oyuncaklar > Bilim Setleri
    "defter":     (3859, 0.18),   # Ofis & Kirtasiye > Okul Urunleri > Okul Defteri
    "lehim":      (2883, 0.16),   # El Aletleri > Lehim ve Havya > Aletleri
    "lehimaks":   (2884, 0.16),   # ... > Lehim ve Havya Aksesuarlari
    "yankeski":   (2859, 0.16),
    "pense":      (2861, 0.16),
    "kargaburun": (2866, 0.16),
    "avadanlik":  (3029, 0.16),   # Takim Cantalari ve Avadanliklar > Avadanliklar
    "takimcanta": (3030, 0.16),
    "multimetre": (3122, 0.13),
    "kontrolkal": (3123, 0.13),
}
OZEL = {
    "AELHMPMP": "lehimaks", "AEHVYSHP": "lehimaks",
    "AEZD30C40W": "lehim", "AEHVYST6P": "lehim",
    "AEMYNK3": "yankeski", "AEYNKSK160": "yankeski",
    "AEPNS180": "pense", "AEEAS3P": "pense", "AEEAS7P": "avadanlik",
    "AEKB1602": "kargaburun",
    "AEDT830D": "multimetre", "AEUT12D": "kontrolkal",
    "AETC162": "takimcanta", "AEPLSTKKT": "takimcanta",
    "AETCS9P": "avadanlik", "AETCS18P": "avadanlik",
}
DEFTER = {"AESTJDFTR", "AEISDSYS10", "AEISDSYS20", "AETMRNDFTR10",
          "AETMRNDFTR20", "AETMRNDFTR30", "AETEMDEF"}

KARGO_PTT = [(0.5, 77.50), (1, 78.75), (2, 81.25), (3, 82.50), (4, 85.00), (5, 86.25),
             (6, 90.00), (7, 103.75), (8, 125.00), (9, 131.25), (10, 138.75), (11, 143.75),
             (12, 153.75), (13, 158.75), (14, 165.00), (15, 153.75), (16, 156.25), (17, 157.50),
             (18, 168.75), (19, 168.75), (20, 168.75), (25, 169.90), (30, 194.90)]


def kargo_ty(t):
    return 34.2 if t < 200 else (65.8 if t < 400 else 77.5)


def kargo_ptt(p, desi):
    if p < 250:
        return 0.0
    for ust, uc in KARGO_PTT:
        if desi <= ust:
            return uc
    return 199.90


def fiyat(p_ty, kom, desi):
    ty_net = p_ty * (1 / KDV - 0.13) - kargo_ty(p_ty) - 11
    p = p_ty
    for _ in range(100):
        yeni = math.ceil((ty_net + kargo_ptt(p, desi)) / (1 / KDV - kom))
        if yeni == p:
            break
        p = yeni
    # 250 TL altinda kargoyu alici odedigi icin esit-net formulu TY'nin cok altina
    # iniyor (hatta eksiye). Kendi kanalimizi baltalamamak icin taban = TY fiyati.
    return max(p, math.ceil(p_ty))


def ean13(gov12):
    s = sum(int(d) * (3 if i % 2 else 1) for i, d in enumerate(gov12))
    return gov12 + str((10 - s % 10) % 10)


def main():
    n11 = {(u.get("tyStokKodu") or u["stokKodu"]): u for u in json.load(
        open(os.path.join(KOK, "content", "n11_urunler.json"), encoding="utf-8"))}
    kat = {u.get("stockCode"): u for u in katalog()}
    pz = {x["sk"]: x for x in json.load(
        open(os.path.join(KOK, "content", "pazarama_fiyat.json"), encoding="utf-8"))}
    hedef = sorted(set(pz) | {"AEPYAA2", "AEPY18650", "AEPY18650T5", "AEPY18650T10", "AEEAS7P"})

    sablon = r"C:\Users\serdar\Downloads\pttavm\pttavm-ornek.xlsx"
    wb = openpyxl.load_workbook(sablon)
    ws = wb["Ürünler"]
    ws.delete_rows(2, ws.max_row)
    for ad in ("Varyant", "Parçalar", "Güvenlik Görselleri&Belgeleri", "Güvenlik Firmaları"):
        s = wb[ad]
        if s.max_row > 1:
            s.delete_rows(2, s.max_row)

    rapor = []
    for i, sk in enumerate(hedef):
        u = kat.get(sk)
        if not u:
            rapor.append((sk, "TY katalogunda yok")); continue
        p_ty = float(u.get("salePrice") or 0)
        if p_ty <= 0:
            rapor.append((sk, "TY fiyati yok")); continue
        anah = OZEL.get(sk) or ("defter" if sk in DEFTER else "bilim")
        kat_id, kom = KAT[anah]
        desi = float((n11.get(sk) or {}).get("desi") or 1)
        p = fiyat(p_ty, kom, desi)
        bar = str((n11.get(sk) or {}).get("barkod") or "")
        if not re.fullmatch(r"\d{13}", bar):
            bar = ean13("29" + str(90000 + i).zfill(10))
        gor = [im["url"] for im in u.get("images", [])][:12]
        gor += [None] * (12 - len(gor))
        acik = _duz_metin(u.get("description") or "") or u["title"]
        satir = [kat_id, bar, sk, u["title"], acik, "Atölye Elektronik",
                 int(u.get("quantity") or 10), p, int(u.get("vatRate") or 20), desi] + gor + [
                 1, 1, None, 1, 0, None, None, 0, 1 if p >= 250 else 0]
        ws.append(satir)
        rapor.append((sk, "%s kat=%d kom=%.0f%% desi=%g TY=%.0f -> PTT=%d" % (anah, kat_id, kom * 100, desi, p_ty, p)))

    cikti = r"C:\Users\serdar\Downloads\pttavm\PTTAVM-urunler.xlsx"
    wb.save(cikti)
    json.dump([{"sk": sk, "not": n} for sk, n in rapor], open(
        os.path.join(KOK, "content", "pttavm_fiyat_rapor.json"), "w", encoding="utf-8"),
        ensure_ascii=False, indent=1)
    ok = sum(1 for _, n in rapor if "->" in n)
    print("yazildi: %s | urun: %d | atlanan: %d" % (cikti, ok, len(rapor) - ok))
    for sk, n in rapor:
        print("  %-14s %s" % (sk, n))


if __name__ == "__main__":
    main()

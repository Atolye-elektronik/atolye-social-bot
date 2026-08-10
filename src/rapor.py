"""Toplanan tüm hedef listelerini tek bir Excel dosyasında birleştirir.

Kaynak CSV'ler ayrı betiklerle üretiliyor (okul_listesi, okul_iletisim,
okul_bolum, ozel_kurumlar) ve `.gitignore` içinde. Bu betik onları okuyup
sekmeli tek bir çalışma kitabı çıkarır; hangi dosya eksikse o sekme atlanır,
böylece tarama sürerken de çalıştırılabilir.

Kullanımı:
    python -m src.rapor
    python -m src.rapor --cikti pazarlama/atolye-hedefler.xlsx
"""

from __future__ import annotations

import argparse
import csv
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PAZARLAMA = pathlib.Path("pazarlama")

KOYU = "245043"
ACIK = "DDE9E4"
BAKIR = "F4E7DA"

# sekme adı -> (dosya, sütunlar, süzgeç)
KURUM_SUTUN = ["il", "ilce", "okul", "eposta", "telefon", "instagram", "bolumler", "adres", "site"]
KURUM_ETIKET = {
    "il": "İl", "ilce": "İlçe", "okul": "Kurum Adı", "eposta": "E-posta", "telefon": "Telefon",
    "instagram": "Instagram", "bolumler": "Alanlar (bölümler)", "adres": "Adres", "site": "Web Sitesi",
}
KURUM_GENISLIK = [14, 18, 50, 22, 16, 20, 56, 44, 34]

OZEL_SUTUN = ["il", "ilce", "kurum", "tur", "telefon", "adres"]
OZEL_ETIKET = {
    "il": "İl", "ilce": "İlçe", "kurum": "Kurum Adı", "tur": "Tür",
    "telefon": "Telefon", "adres": "Adres",
}
OZEL_GENISLIK = [14, 18, 58, 34, 18, 52]

IG_SUTUN = ["segment", "il", "ad", "hesap", "takipci", "iletisim", "not"]
IG_ETIKET = {
    "segment": "Segment", "il": "İl / Bölge", "ad": "Kurum", "hesap": "Instagram",
    "takipci": "Takipçi", "iletisim": "Telefon / E-posta", "not": "Not",
}
IG_GENISLIK = [22, 16, 46, 32, 11, 36, 66]


def oku(ad: str) -> list[dict]:
    yol = PAZARLAMA / ad
    if not yol.exists():
        return []
    with yol.open(encoding="utf-8", newline="") as dosya:
        return list(csv.DictReader(dosya))


def sayfa_ekle(wb: Workbook, baslik: str, satirlar: list[dict],
               sutunlar: list[str], etiketler: dict, genislik: list[int],
               ig_sutun: int | None = None, vurgu: str | None = None) -> int:
    if not satirlar:
        return 0

    ws = wb.create_sheet(baslik[:31])
    ws.append([etiketler[s] for s in sutunlar])
    for hucre in ws[1]:
        hucre.font = Font(bold=True, color="FFFFFF")
        hucre.fill = PatternFill("solid", fgColor=KOYU)
        hucre.alignment = Alignment(vertical="center")

    for satir in satirlar:
        ws.append([satir.get(s, "") for s in sutunlar])
        sira = ws.max_row
        # Takipçi sayısı metin değil sayı olsun ki sıralanabilsin.
        if "takipci" in sutunlar:
            yer = sutunlar.index("takipci") + 1
            deger = str(ws.cell(sira, yer).value or "")
            if deger.isdigit():
                ws.cell(sira, yer).value = int(deger)
        # Instagram hücresini tıklanabilir yap.
        if ig_sutun:
            hucre = ws.cell(sira, ig_sutun)
            if hucre.value:
                hucre.hyperlink = "https://www.instagram.com/" + str(hucre.value).lstrip("@") + "/"
                hucre.font = Font(color="0563C1", underline="single")
        if vurgu and satir.get("elektrik_elektronik") == "var":
            ws.cell(sira, 1).fill = PatternFill("solid", fgColor=vurgu)

    for i, gen in enumerate(genislik, start=1):
        ws.column_dimensions[get_column_letter(i)].width = gen
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(sutunlar))}{ws.max_row}"
    return ws.max_row - 1


def ozet_sayfasi(wb: Workbook, kayitlar: list[tuple[str, int, str]]) -> None:
    ws = wb.create_sheet("ÖZET", 0)
    ws.append(["Sekme", "Kayıt", "Ne işe yarar"])
    for hucre in ws[1]:
        hucre.font = Font(bold=True, color="FFFFFF")
        hucre.fill = PatternFill("solid", fgColor=KOYU)
        hucre.alignment = Alignment(vertical="center")

    for ad, adet, aciklama in kayitlar:
        ws.append([ad, adet, aciklama])
        if adet:
            ws.cell(ws.max_row, 2).fill = PatternFill("solid", fgColor=ACIK)

    for i, gen in enumerate([30, 10, 96], start=1):
        ws.column_dimensions[get_column_letter(i)].width = gen
    for satir in ws.iter_rows(min_row=2):
        satir[2].alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def uret(cikti: pathlib.Path) -> None:
    okullar = oku("okullar.csv")
    ig = oku("instagram-hedefler.csv")

    def sirala(xs):
        return sorted(xs, key=lambda x: (x.get("il", ""), x.get("ilce", ""), x.get("okul", "")))

    def ig_sec(segment):
        veri = [x for x in ig if x.get("segment") == segment]
        return sorted(veri, key=lambda x: (-(int(x["takipci"]) if str(x["takipci"]).isdigit() else -1),
                                           x.get("il", ""), x.get("ad", "")))

    wb = Workbook()
    wb.remove(wb.active)
    ozet: list[tuple[str, int, str]] = []

    def ekle(ad, satirlar, tur="kurum", aciklama="", ig_sut=None, vurgu=None):
        if tur == "kurum":
            n = sayfa_ekle(wb, ad, satirlar, KURUM_SUTUN, KURUM_ETIKET, KURUM_GENISLIK, 6, vurgu)
        elif tur == "ozel":
            n = sayfa_ekle(wb, ad, satirlar, OZEL_SUTUN, OZEL_ETIKET, OZEL_GENISLIK)
        else:
            n = sayfa_ekle(wb, ad, satirlar, IG_SUTUN, IG_ETIKET, IG_GENISLIK, ig_sut)
        if n:
            ozet.append((ad, n, aciklama))

    ee = [x for x in okullar if x.get("elektrik_elektronik") == "var"]
    bt = [x for x in okullar if x.get("bilisim") == "var"]
    hedef = [x for x in okullar if x.get("elektrik_elektronik") == "var" or x.get("bilisim") == "var"]

    ekle("MTAL — Elektrik-Elektronik", sirala(ee), aciklama=(
        "Takım çantası, el aleti ve Arduino setinin asıl alıcısı. Alan bilgisi okulun "
        "kendi site haritasından doğrulandı."))
    ekle("MTAL — Bilişim", sirala(bt), aciklama=(
        "Arduino ve sensör setleri için ikinci hedef."))
    ekle("MTAL — hedef (EE+Bilişim)", sirala(hedef), aciklama=(
        "E-posta kampanyasının ana listesi. okul_daveti.py bu listeyle çalıştırılır."))
    ekle("MTAL — tümü", sirala(okullar), aciklama=(
        "81 ilin tamamı. Alanı 'bilinmiyor' olanlar bölümü yok demek değil — sitesinde "
        "yayınlamamış demek; ikinci dalga için burada."))
    ekle("MESEM", sirala(oku("mesem.csv")), aciklama=(
        "Staj defteri (iş dosyası) hedefi. Öğrencilerin tamamı işletmede eğitim gördüğü "
        "için bölüm filtresi gerekmiyor; hepsi alıcı."))
    ekle("BİLSEM", sirala(oku("bilsem.csv")), aciklama=(
        "Bilim ve Sanat Merkezleri. Robotik/kodlama atölyesi işletiyorlar, bütçeleri var."))
    ekle("Özel MTAL", sirala(oku("ozel-mtal.csv")), tur="ozel", aciklama=(
        "Özel meslek liseleri. Bu dizin e-posta ve site vermiyor — telefon ve adres var."))
    ekle("Özel robotik kursları", sirala(oku("ozel-robotik.csv")), tur="ozel", aciklama=(
        "MEB'e kayıtlı özel robotik/kodlama/bilişim kursları. Yıl boyu malzeme alıyorlar."))

    ekle("IG — Robotik-STEM kursları", ig_sec("Robotik/STEM kursu"), tur="ig", ig_sut=4, aciklama=(
        "Takipçi sayısına göre sıralı. Sezon dışı satışın anahtarı: okullar eylülde alır, "
        "kurslar yıl boyu tüketir."))
    ekle("IG — BİLSEM", ig_sec("BİLSEM"), tur="ig", ig_sut=4)
    ekle("IG — Meslek Yüksekokulu", ig_sec("Meslek Yüksekokulu"), tur="ig", ig_sut=4)
    ekle("IG — Meslek Liseleri", ig_sec("Meslek Lisesi (MTAL)"), tur="ig", ig_sut=4, aciklama=(
        "DM'e elektrik-elektronik teyitli olanlardan başla."))
    ekle("WhatsApp (cep no)", oku("okul-whatsapp.csv"), aciklama=(
        "Okul telefonlarının %95'i sabit hat ve WhatsApp'a kapalı; bunlar cep numarası "
        "olan okullar."))

    ozet_sayfasi(wb, ozet)
    cikti.parent.mkdir(parents=True, exist_ok=True)
    wb.save(cikti)

    print(f"{cikti} yazıldı:")
    for ad, adet, _ in ozet:
        print(f"  {ad:<30} {adet:>5}")
    print(f"  {'TOPLAM':<30} {sum(a for _, a, _ in ozet):>5}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Tüm hedef listelerini tek Excel'de topla")
    parser.add_argument("--cikti", type=pathlib.Path,
                        default=PAZARLAMA / "atolye-hedefler.xlsx")
    args = parser.parse_args()
    uret(args.cikti)


if __name__ == "__main__":
    main()
